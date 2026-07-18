"""
Import the barrel/rack seed (the template we specified) into the physical fleet.

    python manage.py import_barrel_seed path/to/barrels.xlsx [--wipe] [--dry-run]

One row per barrel. Creates/updates Room, Location (room + column), Rack (with
size_class), Container (barrel or SS drum, with pool), and the current
RackAssignment. For a filled barrel it opens an AgingPlacement carrying the
historical wine label in `legacy_lot_code` (occupancy without a Lot FK — the
Option-A representation); `EMPTY` means an empty barrel (no placement).

Integrity guarantees:
  * `fill_number` is set EXPLICITLY from `prior_fills` (never the count-based
    default), so importing this now and real 2023/24 harvest lots later can't
    corrupt oak tiers regardless of order.
  * barrel_id is read as text (leading zeros preserved when the source file
    stores them as text) and only needs to be unique + numeric.
  * a rack whose two rows disagree on location is assigned ONE location from its
    position-1 row and reported.
  * SS drums (55 gal) import as type=ss_drum and are excluded from oak tiers
    (Container.is_oak is False for them).

Idempotent: re-running updates existing rows (matched on container_id / rack_id)
rather than duplicating. `--wipe` clears the fleet first for a clean reload.
"""
from collections import Counter, defaultdict
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from cellar.models import (
    Container, Rack, Room, Location, RackAssignment, AgingPlacement,
)

HEADSPACE = Decimal("3")
ROOM_ABBR = {
    "old barrel room": "OBC",
    "new barrel room": "NBC",
    "bldg. h": "BLDGH", "bldg h": "BLDGH", "building h": "BLDGH",
}
EMPTY = "EMPTY"


def _abbr(room_name):
    key = (room_name or "").strip().lower()
    if key in ROOM_ABBR:
        return ROOM_ABBR[key]
    return "".join(w[0] for w in key.split()).upper() or "RM"


def _norm_pool(v):
    v = (str(v or "").strip().lower())
    return v if v in ("table", "port") else ""


def _size_class(size):
    return "large" if int(size) == 130 else "standard"


def _ctype(size):
    return Container.Type.SS_DRUM if int(size) == 55 else Container.Type.BARREL


class Command(BaseCommand):
    help = "Import the barrel/rack seed spreadsheet into the physical fleet."

    def add_arguments(self, parser):
        parser.add_argument("path")
        parser.add_argument("--wipe", action="store_true",
                            help="clear Container/Rack/RackAssignment/legacy placements first")
        parser.add_argument("--dry-run", action="store_true",
                            help="validate and report, write nothing")

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("openpyxl is required: pip install openpyxl")

        wb = load_workbook(opts["path"], data_only=True)
        ws = wb[wb.sheetnames[0]]
        hdr = None
        rows = []
        for r in ws.iter_rows(values_only=True):
            if hdr is None:
                hdr = [str(c).strip() if c is not None else c for c in r]
                continue
            if r[0] is None:
                continue
            rows.append(dict(zip(hdr, r)))

        report = _Report()
        parsed = self._validate(rows, report)
        report.emit(self.stdout, self.style)
        if report.fatal:
            raise CommandError("Import aborted — fix the fatal issues above.")
        if opts["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry run — nothing written."))
            return

        with transaction.atomic():
            if opts["wipe"]:
                AgingPlacement.objects.filter(lot__isnull=True).delete()
                RackAssignment.objects.all().delete()
                Rack.objects.all().delete()
                Container.objects.all().delete()
                self.stdout.write(self.style.WARNING("Wiped existing fleet."))
            self._write(parsed, report)

        self.stdout.write(self.style.SUCCESS(
            f"Imported {report.counts['containers']} barrels on "
            f"{report.counts['racks']} racks across {report.counts['locations']} locations."))

    # ------------------------------------------------------------------
    def _validate(self, rows, report):
        seen_ids = set()
        by_rack = defaultdict(list)
        parsed = []
        for i, r in enumerate(rows, start=2):
            bid = r.get("barrel_id")
            if bid is None:
                report.fatal_msg(f"row {i}: missing barrel_id")
                continue
            bid = str(bid).strip()
            if not bid.isdigit():
                report.warn("non-numeric barrel_id", bid)
            if bid in seen_ids:
                report.fatal_msg(f"row {i}: duplicate barrel_id {bid}")
                continue
            seen_ids.add(bid)

            try:
                size = int(r.get("size_gal"))
            except (TypeError, ValueError):
                report.fatal_msg(f"row {i}: bad size_gal {r.get('size_gal')!r}")
                continue
            if size not in (55, 60, 70, 130):
                report.warn("unexpected size_gal", size)

            pool = _norm_pool(r.get("pool"))
            room = (str(r.get("room") or "").strip())
            col = r.get("column")
            col = int(col) if isinstance(col, (int, float)) and str(col).strip() != "" else None
            if col is None:
                report.note("room-only (no column)", f"{bid} in {room or '?'}")
            rack_id = r.get("rack_id")
            rack_id = str(rack_id).strip() if rack_id not in (None, "") else None
            pos = r.get("rack_position")
            pos = int(pos) if pos not in (None, "") else None
            lot_code = str(r.get("current_lot") or "").strip()
            filled = bool(lot_code) and lot_code.upper() != EMPTY
            fill_date = r.get("fill_date")
            if hasattr(fill_date, "date"):
                fill_date = fill_date.date()
            prior = r.get("prior_fills")
            prior = int(prior) if prior not in (None, "") else None

            rec = dict(bid=bid, size=size, pool=pool, room=room, col=col,
                       rack_id=rack_id, pos=pos, filled=filled,
                       lot_code=lot_code if filled else "", fill_date=fill_date,
                       prior=prior, barcode=str(r.get("barrel_barcode") or "").strip())
            parsed.append(rec)
            if rack_id:
                by_rack[rack_id].append(rec)

        # rack-level integrity + one-location resolution
        rack_loc = {}
        for rid, members in by_rack.items():
            if len(members) > 2:
                report.fatal_msg(f"rack {rid}: {len(members)} barrels (max 2)")
            positions = [m["pos"] for m in members]
            if len(set(positions)) != len(positions):
                report.fatal_msg(f"rack {rid}: duplicate positions {positions}")
            pools = {m["pool"] for m in members if m["pool"]}
            if len(pools) > 1:
                report.fatal_msg(f"rack {rid}: mixed pool {pools}")
            classes = {_size_class(m["size"]) for m in members}
            if len(classes) > 1:
                report.fatal_msg(f"rack {rid}: mixed size-class {classes}")
            # location: prefer position 1, else first row
            anchor = min(members, key=lambda m: (m["pos"] is None, m["pos"]))
            rooms = {m["room"] for m in members}
            cols = {m["col"] for m in members}
            if len(rooms) > 1 or len(cols) > 1:
                report.conflict(rid, members, anchor)
            rack_loc[rid] = (anchor["room"], anchor["col"],
                             "large" if "large" in classes else "standard")
        self._rack_loc = rack_loc
        return parsed

    # ------------------------------------------------------------------
    def _write(self, parsed, report):
        rooms = {}
        locations = {}

        def get_room(name):
            key = name.strip().lower()
            if key not in rooms:
                rooms[key], _ = Room.objects.get_or_create(
                    name=name.strip() or "Unspecified", defaults={"notes": ""})
            return rooms[key]

        def get_location(room_name, col):
            code = f"{_abbr(room_name)}{col}" if col is not None else _abbr(room_name)
            if code not in locations:
                room = get_room(room_name)
                loc, _ = Location.objects.get_or_create(
                    code=code, defaults={"room": room})
                if loc.room_id != room.id:
                    loc.room = room
                    loc.save(update_fields=["room"])
                locations[code] = loc
            return locations[code]

        # racks first (so assignments can attach)
        racks = {}
        for rid, (room_name, col, sclass) in self._rack_loc.items():
            loc = get_location(room_name, col)
            rack, _ = Rack.objects.get_or_create(rack_id=rid)
            rack.location = loc
            rack.size_class = sclass
            rack.save()
            racks[rid] = rack
        report.counts["racks"] = len(racks)
        report.counts["locations"] = len(locations)

        now = timezone.now()
        made_c = 0
        for rec in parsed:
            c, _ = Container.objects.get_or_create(
                container_id=rec["bid"],
                defaults={"type": _ctype(rec["size"]),
                          "capacity_gal": Decimal(str(rec["size"]))})
            c.type = _ctype(rec["size"])
            c.capacity_gal = Decimal(str(rec["size"]))
            c.pool = rec["pool"]
            c.format = f"{rec['size']} gal"
            if rec["barcode"]:
                c.barcode = rec["barcode"]
            c.active = True
            c.save()
            made_c += 1
            report.by_type[c.get_type_display()] += 1
            report.by_pool[rec["pool"] or "—"] += 1
            report.by_size[rec["size"]] += 1

            # rack assignment (current)
            if rec["rack_id"] and rec["pos"] is not None:
                rack = racks.get(rec["rack_id"])
                if rack and not RackAssignment.objects.filter(
                        container=c, removed_at__isnull=True).exists():
                    RackAssignment.objects.create(
                        container=c, rack=rack, position=rec["pos"], assigned_at=now)

            # occupancy: legacy placement for a filled barrel
            if rec["filled"]:
                if not AgingPlacement.objects.filter(
                        container=c, emptied_at__isnull=True, voided_at__isnull=True).exists():
                    fill_number = (rec["prior"] + 1) if rec["prior"] is not None else 1
                    cap = Decimal(str(rec["size"]))
                    vol = (cap - HEADSPACE) if cap > HEADSPACE else cap
                    AgingPlacement.objects.create(
                        container=c, lot=None, legacy_lot_code=rec["lot_code"],
                        filled_at=rec["fill_date"] or timezone.localdate(),
                        volume_gal=vol.quantize(Decimal("0.1")),
                        fill_number=fill_number)  # oak_tier auto-derives from this
                    report.counts["filled"] += 1
            else:
                report.counts["empty"] += 1
        report.counts["containers"] = made_c


class _Report:
    def __init__(self):
        self.fatal = False
        self._fatal = []
        self.warns = Counter()
        self._warn_ex = defaultdict(list)
        self.notes = Counter()
        self._note_ex = defaultdict(list)
        self.conflicts = []
        self.counts = Counter()
        self.by_type = Counter()
        self.by_pool = Counter()
        self.by_size = Counter()

    def fatal_msg(self, m):
        self.fatal = True
        self._fatal.append(m)

    def warn(self, kind, ex):
        self.warns[kind] += 1
        if len(self._warn_ex[kind]) < 6:
            self._warn_ex[kind].append(str(ex))

    def note(self, kind, ex):
        self.notes[kind] += 1
        if len(self._note_ex[kind]) < 6:
            self._note_ex[kind].append(str(ex))

    def conflict(self, rid, members, anchor):
        self.conflicts.append(
            (rid, [(m["bid"], m["room"], m["col"], m["pos"]) for m in members],
             (anchor["room"], anchor["col"])))

    def emit(self, out, style):
        out.write(style.MIGRATE_HEADING("\n=== Barrel seed report ==="))
        if self._fatal:
            out.write(style.ERROR(f"FATAL ({len(self._fatal)}):"))
            for m in self._fatal[:20]:
                out.write(style.ERROR(f"  • {m}"))
        for kind, n in self.warns.items():
            out.write(style.WARNING(f"WARN  {kind}: {n}  e.g. {self._warn_ex[kind]}"))
        for kind, n in self.notes.items():
            out.write(f"note  {kind}: {n}  e.g. {self._note_ex[kind]}")
        if self.conflicts:
            out.write(style.WARNING(
                f"\nLocation conflicts ({len(self.conflicts)}) — assigned from position-1 barrel:"))
            for rid, members, chosen in self.conflicts:
                out.write(style.WARNING(f"  {rid}: {members}  → kept {chosen}"))
        out.write("")
